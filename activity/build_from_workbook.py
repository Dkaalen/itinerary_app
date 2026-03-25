from __future__ import annotations

import json
import math
import re
from collections import defaultdict
from datetime import datetime, time
from pathlib import Path

import openpyxl


COUNTRY_MAP = {
    "NO": "Norway",
    "FI": "Finland",
    "IS": "Iceland",
    "SE": "Sweden",
    "DK": "Denmark",
    "Resorts": "Resorts",
    "Extras": "Extras",
}
SHEETS = ["NO", "FI", "IS", "SE", "DK", "Resorts", "Extras"]

MARKER_ALIASES = {
    "Time": ["Time", "Times"],
    "Meeting point": ["Meeting point", "Meeting Point"],
    "Includes": ["Includes", "Include"],
    "Notable sights": ["Notable sights", "Notable Sights"],
    "End point": ["End point", "Endpoint"],
    "Departure from": ["Departure from"],
    "Return from": ["Return from"],
    "Duration": ["Duration"],
    "Pick-up": ["Pick-up"],
    "Drop-off": ["Drop-off"],
    "Note": ["Note", "Notes"],
    "Level": ["Level"],
    "Minimum age": ["Minimum age"],
}
MARKER_PATTERNS: list[str] = []
for variants in MARKER_ALIASES.values():
    for variant in variants:
        MARKER_PATTERNS.append(re.escape(variant))


def slugify(value: str) -> str:
    value = (value or "").strip().lower()
    replacements = {
        "å": "a", "ø": "o", "æ": "ae", "ö": "o", "ä": "a",
        "í": "i", "ð": "d", "þ": "th", "ú": "u", "ó": "o",
        "é": "e", "á": "a", "ñ": "n", "ý": "y", "ü": "u", "ß": "ss",
    }
    for old, new in replacements.items():
        value = value.replace(old, new)
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-")


def split_city_and_rest(text: str) -> tuple[str | None, str]:
    text = (text or "").strip()
    if ":" in text:
        left, right = text.split(":", 1)
        if len(left) <= 40:
            return left.strip(), right.strip()
    return None, text


def parse_details(rest: str) -> tuple[str, dict[str, str]]:
    details: dict[str, str] = {}
    title = rest.strip()
    pattern = re.compile(r"\s+-\s+(" + "|".join(MARKER_PATTERNS) + r")\s*:\s*", flags=re.I)
    match = pattern.search(rest)
    if not match:
        return title, details

    title = rest[:match.start()].strip(" -")
    parts = pattern.split(rest[match.start():])

    index = 1
    while index < len(parts):
        raw_key = parts[index].strip()
        value = parts[index + 1] if index + 1 < len(parts) else ""
        canonical = next(
            (name for name, variants in MARKER_ALIASES.items() if raw_key.lower() in [v.lower() for v in variants]),
            raw_key,
        )
        details[canonical] = value.strip(" -")
        index += 2

    return title, details


def normalize_value(value):
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, float):
        if math.isnan(value):
            return None
        return round(value, 2)
    return value


def build_catalog(workbook_path: str | Path, output_path: str | Path) -> None:
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    headers = [workbook["NO"].cell(6, column).value for column in range(2, 37)]
    column_map = {header: idx + 2 for idx, header in enumerate(headers)}

    records: list[dict] = []
    seen: set[tuple] = set()

    for sheet_name in SHEETS:
        sheet = workbook[sheet_name]
        for row_index in range(7, sheet.max_row + 1):
            row = {header: sheet.cell(row_index, column_map[header]).value for header in headers}
            if row.get("Type") != "Activity" or not row.get("Travel element"):
                continue

            raw = str(row["Travel element"]).strip()
            city_prefix, rest = split_city_and_rest(raw)
            title, details = parse_details(rest)
            destination = city_prefix or row.get("ID") or "Unknown"

            record = {
                "source_sheet": sheet_name,
                "country": COUNTRY_MAP.get(sheet_name, sheet_name),
                "destination": destination,
                "destination_id": row.get("ID") or destination,
                "slug": slugify(title or raw),
                "title": title or raw,
                "raw_text": raw,
                "type": row.get("Type"),
                "price_eur_adult": normalize_value(row.get("Sales P per unit")),
                "price_currency": row.get("Sales curr") or "EUR",
                "supplier": row.get("Supplier"),
                "supplier_url": row.get("URL"),
                "status": row.get("Status"),
                "comments": row.get("Comments"),
                "booking_notes": {
                    "manual_booking": bool(row.get("Manual booking?")),
                    "non_refundable": bool(row.get("Non-refundable")),
                    "refundable": bool(row.get("Refundable")),
                },
                "details": {key: value for key, value in details.items() if value not in (None, "")},
                "spreadsheet_row": row_index,
            }

            dedupe_key = (
                record["destination"],
                record["title"],
                record["supplier_url"],
                record["price_eur_adult"],
            )
            if dedupe_key in seen:
                continue

            seen.add(dedupe_key)
            records.append(record)

    records.sort(key=lambda item: (item["country"], item["destination"], item["title"]))

    payload = {
        "generated_from": Path(workbook_path).name,
        "activity_count": len(records),
        "destinations": sorted({item["destination"] for item in records}),
        "activities": records,
    }

    with Path(output_path).open("w", encoding="utf-8") as file:
        json.dump(payload, file, ensure_ascii=False, indent=2)


if __name__ == "__main__":
    root = Path(__file__).resolve().parent
    build_catalog(root.parent / "Cheat Sheet 2.0.xlsx", root / "data.json")